# ==============================================================================
# modules
# ==============================================================================
import os
import sys
from L3Out import ModularL3Out
import openpyxl
# ==============================================================================

# ==============================================================================
# changeDir
# ==============================================================================
if os.path.dirname(sys.argv[0]) != "":
    os.chdir(os.path.dirname(sys.argv[0]))

# ==============================================================================
# check excel-content
# ==============================================================================
def check_set(content):
    if content != "None" and content != "" and content is not None:
        return True
    else:
        return False


def get_excel_data():
    xlsxFo = "L3OUT_eBGP.xlsx"
    wbObj = openpyxl.load_workbook(xlsxFo, data_only=True)
    sheet = wbObj.get_sheet_by_name("L3OUT")

    L3Out = []
    Keys = []

    for row in range(2, sheet.max_row + 1):
        if row == 2:
            for col in range(1,sheet.max_column+1):
                value = sheet.cell(row=row, column=col).value
                if value != None: Keys.append(value)
            # print(Keys)
        else:
            if check_set(str(sheet["A" + str(row)].value)):
                line = {}
                for x in range(len(Keys)):
                    line[Keys[x]]=str(sheet.cell(row=row, column=x+1).value)
                L3Out.append(line)
    return L3Out


def get_remote_bgp_ip(ip):
    BGP_IP = ip.partition(("/")[0])[0].split(".")
    BGP_IP[3] = int(BGP_IP[3]) - 1
    return '.'.join(str(e) for e in BGP_IP)


if __name__ == '__main__':

    Tenant = "TN-PROD"
    Contract = "CT-PERMIT-ALL"
    Path1 = "topology/pod-1/node-111"
    Paths1 = "topology/pod-1/paths-111/pathep-[eth1/50]"
    Paths1MTier = "topology/pod-1/paths-111/pathep-[eth1/49]"

    Path2 = "topology/pod-1/node-112"
    Paths2 = "topology/pod-1/paths-112/pathep-[eth1/50]"
    Paths2MTier = "topology/pod-1/paths-112/pathep-[eth1/49]"

    Path3 = "topology/pod-1/node-211"
    Paths3 = "topology/pod-1/paths-211/pathep-[eth1/50]"
    Paths3MTier = "topology/pod-1/paths-211/pathep-[eth1/49]"

    Path4 = "topology/pod-1/node-212"
    Paths4 = "topology/pod-1/paths-212/pathep-[eth1/50]"
    Paths4MTier = "topology/pod-1/paths-212/pathep-[eth1/49]"

    Lo1 = "10.10.20.128"
    Lo2 = "10.10.20.129"
    Lo3 = "10.10.20.142"
    Lo4 = "10.10.20.143"

    BGP_PWD = "SecureTest"
    AST1 = "65164"
    LocalAS = "65162"

    L3Out_Data = get_excel_data()

    for item in L3Out_Data:

        Tenant = "TN-PROD"

        L3Out2Post = ModularL3Out.L3Out(item["NAME"], Tenant)
        L3Out2Post.setl3domain(item["Dom"])
        L3Out2Post.setVrf(item["VRF"])
        L3Out2Post.setExternalEpg(item["Ext-EPG"].split(":")[0])
        L3Out2Post.externalEpg().setConsumeContract(Contract)
        L3Out2Post.externalEpg().setProvideContract(Contract)
        L3Out2Post.externalEpg().setL3ExtSubnet()
        L3Out2Post.setEnableBgp()


        L3Out2Post.setNodeProfile(item["NP"])
        L3Out2Post.nodeProfile().setNode(Path1, Lo1)
        L3Out2Post.nodeProfile().setNode(Path2, Lo2)
        L3Out2Post.nodeProfile().setNode(Path3, Lo3)
        L3Out2Post.nodeProfile().setNode(Path4, Lo4)

        L3Out2Post.nodeProfile().setInt(item["INT"])
        L3Out2Post.nodeProfile().Int().setIntNode(item["TYP"], "vlan-" + item["VLAN1"], item["P-IP1"] + "/30", Paths1)

        RemoteIP = get_remote_bgp_ip(item["P-IP1"])

        L3Out2Post.nodeProfile().Int().intNode().setBgpPeer(RemoteIP, BGP_PWD)
        L3Out2Post.nodeProfile().Int().intNode().bgpPeer().setBgpAS(AST1)
        L3Out2Post.nodeProfile().Int().intNode().bgpPeer().setBgpLocalAS(LocalAS)

        L3Out2Post.nodeProfile().Int().setIntNode(item["TYP"], "vlan-" + item["VLAN2"], item["P-IP2"] + "/30", Paths2)

        RemoteIP = get_remote_bgp_ip(item["P-IP2"])

        L3Out2Post.nodeProfile().Int().intNode().setBgpPeer(RemoteIP, BGP_PWD)
        L3Out2Post.nodeProfile().Int().intNode().bgpPeer().setBgpAS(AST1)

        L3Out2Post.nodeProfile().Int().setIntNode(item["TYP"], "vlan-" + item["VLAN3"], item["P-IP3"] + "/30", Paths3)

        RemoteIP = get_remote_bgp_ip(item["P-IP3"])

        L3Out2Post.nodeProfile().Int().intNode().setBgpPeer(RemoteIP, BGP_PWD)
        L3Out2Post.nodeProfile().Int().intNode().bgpPeer().setBgpAS(AST1)

        L3Out2Post.nodeProfile().Int().setIntNode(item["TYP"], "vlan-" + item["VLAN4"], item["P-IP4"] + "/30", Paths4)

        RemoteIP = get_remote_bgp_ip(item["P-IP4"])

        L3Out2Post.nodeProfile().Int().intNode().setBgpPeer(RemoteIP, BGP_PWD)
        L3Out2Post.nodeProfile().Int().intNode().bgpPeer().setBgpAS(AST1)
        L3Out2Post.nodeProfile().Int().intNode().bgpPeer().setBgpLocalAS(LocalAS)

        print(L3Out2Post.tostring())
